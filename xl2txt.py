# Script to read words and their meanings from wordlist.xlsx and write to vocab.txt sequentially

import openpyxl
wb = openpyxl.load_workbook('wordlist.xlsx')
sht = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sht[0])

# helloFile = open('hello.txt')
# helloContent = helloFile.read()	# read the entire contents of a file as a string value
# helloFile.close()

# helloFile = open('hello.txt','r')	# w, a, w+
# #helloFile.write('Bacon is not a vegetable.')	# does not automatically add a newline character to the end of the string like the print() function does
# hello2 = helloFile.readlines()	# get a list of string values from the file, one string for each line of text
# helloFile.close()

# print(helloContent)
# print(hello2)

vocabFile = open('vocab.txt','w')

for i in range(2, sheet.max_row):
	vocabFile.write(sheet.cell(row=i, column=1).value+'\n')
	vocabFile.write(sheet.cell(row=i, column=2).value+'\n')

vocabFile.close()